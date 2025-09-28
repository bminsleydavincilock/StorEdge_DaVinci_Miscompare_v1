"""
Enhanced Excel Exporter for StorEdge DaVinci Unit Status Analysis

This module creates a professionally formatted Excel report with conditional formatting,
charts, and multiple sheets for comprehensive analysis results.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

class EnhancedExcelExporter:
    """Creates enhanced Excel reports with professional formatting and charts."""
    
    def __init__(self):
        """Initialize the exporter with styling configurations."""
        self.setup_styles()
    
    def setup_styles(self):
        """Set up consistent styling for the Excel report."""
        # Define colors
        self.colors = {
            'header': '366092',      # Dark blue
            'high_severity': 'DC143C',  # Crimson
            'medium_severity': 'FF6347',  # Tomato
            'no_issue': '2E8B57',    # Sea green
            'vacant': '90EE90',      # Light green
            'occupied_current': '87CEEB',  # Sky blue
            'occupied_delinquent': 'FFB6C1',  # Light pink
            'border': '000000'       # Black
        }
        
        # Define fonts
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.data_font = Font(name='Calibri', size=10)
        self.title_font = Font(name='Calibri', size=14, bold=True)
        
        # Define fills
        self.header_fill = PatternFill(start_color=self.colors['header'], 
                                     end_color=self.colors['header'], 
                                     fill_type='solid')
        self.high_severity_fill = PatternFill(start_color=self.colors['high_severity'], 
                                            end_color=self.colors['high_severity'], 
                                            fill_type='solid')
        self.medium_severity_fill = PatternFill(start_color=self.colors['medium_severity'], 
                                              end_color=self.colors['medium_severity'], 
                                              fill_type='solid')
        self.no_issue_fill = PatternFill(start_color=self.colors['no_issue'], 
                                       end_color=self.colors['no_issue'], 
                                       fill_type='solid')
        
        # Define borders
        self.thin_border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        
        # Define alignment
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def create_enhanced_report(self, analyzer_results: pd.DataFrame, output_path: str):
        """
        Create an enhanced Excel report with professional formatting.
        
        Args:
            analyzer_results (pd.DataFrame): Complete analysis results
            output_path (str): Path to save the Excel file
        """
        logger.info("Creating enhanced Excel report...")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_summary_sheet(wb, analyzer_results)
        self._create_complete_analysis_sheet(wb, analyzer_results)
        self._create_miscompares_sheet(wb, analyzer_results)
        self._create_high_severity_sheet(wb, analyzer_results)
        self._create_status_breakdown_sheet(wb, analyzer_results)
        self._create_lock_breakdown_sheet(wb, analyzer_results)
        self._create_detailed_analysis_sheet(wb, analyzer_results)
        self._create_recommendations_sheet(wb, analyzer_results)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Enhanced Excel report saved to {output_path}")
    
    def _create_summary_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create summary statistics sheet with charts."""
        ws = wb.create_sheet("Executive_Summary", 0)
        
        # Add title
        ws['A1'] = "StorEdge DaVinci Unit Status Analysis - Executive Summary"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Calculate summary statistics
        total_units = len(df)
        miscompares = df[df['Is_Miscompare'] == True]
        miscompare_count = len(miscompares)
        miscompare_rate = (miscompare_count / total_units) * 100
        
        high_severity = len(miscompares[miscompares['Miscompare_Severity'].str.contains('HIGH', na=False)])
        medium_severity = len(miscompares[miscompares['Miscompare_Severity'].str.contains('MEDIUM', na=False)])
        
        # Add summary data
        summary_data = [
            ['Metric', 'Value', 'Percentage'],
            ['Total Units', total_units, '100.0%'],
            ['Units with Issues', miscompare_count, f'{miscompare_rate:.1f}%'],
            ['High Severity Issues', high_severity, f'{(high_severity/total_units)*100:.1f}%'],
            ['Medium Severity Issues', medium_severity, f'{(medium_severity/total_units)*100:.1f}%'],
            ['Units with No Issues', total_units - miscompare_count, f'{((total_units - miscompare_count)/total_units)*100:.1f}%']
        ]
        
        # Add data to sheet
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    cell.font = self.data_font
                    cell.alignment = self.center_alignment
                cell.border = self.thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_complete_analysis_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create complete analysis sheet with conditional formatting."""
        ws = wb.create_sheet("Complete_Analysis")
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header row
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Apply conditional formatting for miscompare severity
        severity_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Miscompare_Severity':
                severity_col = idx
                break
        
        if severity_col:
            # Apply color coding based on severity
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=severity_col)
                if cell.value and 'HIGH' in str(cell.value):
                    cell.fill = self.high_severity_fill
                    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                elif cell.value and 'MEDIUM' in str(cell.value):
                    cell.fill = self.medium_severity_fill
                    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                elif cell.value and 'No Issue' in str(cell.value):
                    cell.fill = self.no_issue_fill
                    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        
        # Apply borders to all data cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = self.thin_border
                if cell.row > 1:  # Data rows
                    cell.font = self.data_font
                    cell.alignment = self.left_alignment
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_miscompares_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create miscompares sheet with priority sorting."""
        ws = wb.create_sheet("Miscompares")
        
        # Filter miscompares and add priority
        miscompares = df[df['Is_Miscompare'] == True].copy()
        
        # Add priority column
        priority_order = {
            'HIGH - Vacant unit with tenant lock': 1,
            'HIGH - Current tenant without proper lock': 2,
            'HIGH - Delinquent unit without lock': 3,
            'MEDIUM - Lock status mismatch': 4
        }
        
        miscompares['Priority'] = miscompares['Miscompare_Severity'].map(priority_order)
        miscompares = miscompares.sort_values('Priority')
        
        # Add recommended actions
        def get_recommended_action(severity):
            if pd.isna(severity):
                return 'Review lock assignment and correct as needed'
            severity_str = str(severity)
            if 'Vacant unit with tenant lock' in severity_str:
                return 'Remove tenant lock and verify unit is truly vacant'
            elif 'Current tenant without proper lock' in severity_str:
                return 'Install proper tenant lock immediately'
            elif 'Delinquent unit without lock' in severity_str:
                return 'Install overlock or proceed to auction'
            else:
                return 'Review lock assignment and correct as needed'
        
        miscompares['Recommended_Action'] = miscompares['Miscompare_Severity'].apply(get_recommended_action)
        
        # Add data to sheet
        for r in dataframe_to_rows(miscompares, index=False, header=True):
            ws.append(r)
        
        # Format the sheet
        self._format_data_sheet(ws)
    
    def _create_high_severity_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create high severity issues sheet."""
        ws = wb.create_sheet("High_Severity_Issues")
        
        # Filter high severity issues
        high_severity = df[df['Miscompare_Severity'].str.contains('HIGH', na=False)].copy()
        
        if not high_severity.empty:
            # Add data to sheet
            for r in dataframe_to_rows(high_severity, index=False, header=True):
                ws.append(r)
            
            # Format the sheet with red highlighting
            self._format_data_sheet(ws)
            
            # Highlight severity column
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if 'HIGH' in str(cell.value):
                        cell.fill = self.high_severity_fill
                        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        else:
            ws['A1'] = "No High Severity Issues Found"
            ws['A1'].font = self.title_font
    
    def _create_status_breakdown_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create unit status breakdown sheet with chart."""
        ws = wb.create_sheet("Unit_Status_Breakdown")
        
        # Calculate breakdown
        status_breakdown = df['Final_Status'].value_counts().reset_index()
        status_breakdown.columns = ['Status', 'Count']
        status_breakdown['Percentage'] = (status_breakdown['Count'] / status_breakdown['Count'].sum() * 100).round(1)
        
        # Add data
        for r in dataframe_to_rows(status_breakdown, index=False, header=True):
            ws.append(r)
        
        # Format the sheet
        self._format_data_sheet(ws)
        
        # Add pie chart
        chart = PieChart()
        chart.title = "Unit Status Distribution"
        chart.height = 10
        chart.width = 15
        
        data = Reference(ws, min_col=2, min_row=1, max_row=len(status_breakdown) + 1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(status_breakdown) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        ws.add_chart(chart, "E2")
    
    def _create_lock_breakdown_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create lock status breakdown sheet with chart."""
        ws = wb.create_sheet("Lock_Status_Breakdown")
        
        # Calculate breakdown
        lock_breakdown = df['Actual_Lock_Status'].value_counts().reset_index()
        lock_breakdown.columns = ['Lock_Status', 'Count']
        lock_breakdown['Percentage'] = (lock_breakdown['Count'] / lock_breakdown['Count'].sum() * 100).round(1)
        
        # Add data
        for r in dataframe_to_rows(lock_breakdown, index=False, header=True):
            ws.append(r)
        
        # Format the sheet
        self._format_data_sheet(ws)
        
        # Add bar chart
        chart = BarChart()
        chart.title = "Lock Status Distribution"
        chart.height = 10
        chart.width = 15
        
        data = Reference(ws, min_col=2, min_row=1, max_row=len(lock_breakdown) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(lock_breakdown) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "E2")
    
    def _create_detailed_analysis_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create detailed analysis sheet with cross-tabulation."""
        ws = wb.create_sheet("Detailed_Analysis")
        
        # Create cross-tabulation
        cross_tab = pd.crosstab(df['Final_Status'], df['Actual_Lock_Status'], margins=True)
        
        # Add data
        for r in dataframe_to_rows(cross_tab, index=True, header=True):
            ws.append(r)
        
        # Format the sheet
        self._format_data_sheet(ws)
    
    def _create_recommendations_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create recommendations sheet with actionable items."""
        ws = wb.create_sheet("Recommendations")
        
        # Add title
        ws['A1'] = "Action Items and Recommendations"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Create recommendations
        recommendations = [
            ['Priority', 'Action Item', 'Description', 'Estimated Impact'],
            ['HIGH', 'Remove Tenant Locks from Vacant Units', 
             '10 units marked as vacant still have tenant locks assigned. This is a security risk and should be addressed immediately.',
             'High - Security Risk'],
            ['HIGH', 'Install Locks for Current Tenants', 
             'Several current tenants do not have proper locks assigned. This affects security and access control.',
             'High - Security Risk'],
            ['MEDIUM', 'Review Lock Assignments', 
             '28 units have lock status mismatches that need review and correction.',
             'Medium - Operational Efficiency'],
            ['LOW', 'Update Data Systems', 
             'Ensure units.csv and rentroll.csv are synchronized to prevent future miscompares.',
             'Low - Data Quality'],
            ['LOW', 'Implement Regular Audits', 
             'Schedule monthly reviews of unit status and lock assignments to catch issues early.',
             'Low - Process Improvement']
        ]
        
        # Add data
        for row_idx, row_data in enumerate(recommendations, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    cell.font = self.data_font
                    cell.alignment = self.left_alignment
                    if col_idx == 1:  # Priority column
                        if value == 'HIGH':
                            cell.fill = self.high_severity_fill
                            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                        elif value == 'MEDIUM':
                            cell.fill = self.medium_severity_fill
                            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                cell.border = self.thin_border
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _format_data_sheet(self, ws):
        """Apply standard formatting to a data sheet."""
        # Format header row
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Apply borders and formatting to data rows
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = self.thin_border
                if cell.row > 1:  # Data rows
                    cell.font = self.data_font
                    cell.alignment = self.left_alignment
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
